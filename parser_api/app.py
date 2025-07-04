from flask import Flask, request, jsonify
import requests
import fitz  # PyMuPDF
from docx import Document  # python-docx
from io import BytesIO, StringIO
import openpyxl  # For Excel support
import csv

app = Flask(__name__)
valid_file_types = ["pdf", "epub", "txt", "docx", "xlsx", "csv"]

def get_file_type(url: str):
    path = url.split("?")[0]
    return path.split(".")[-1].lower()

def read_with_PyMuPDF(content, file_type):
    """ Supported file types: PDF, EPUB, TXT """
    pdf_document = fitz.open(stream=content, filetype=file_type)
    extracted_text = "".join(page.get_text() for page in pdf_document)
    return extracted_text

def read_docx(content):
    document = Document(content)
    return "\n".join(paragraph.text for paragraph in document.paragraphs)

def read_xlsx(content):
    workbook = openpyxl.load_workbook(content, data_only=True)
    extracted_text = []
    
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        rows = list(worksheet.iter_rows(values_only=True))
        
        if not rows:
            continue
        
        headers = rows[0]  # First row as headers
        for row in rows[1:]:  # Iterate over remaining rows
            extracted_text.append("("+",\t".join(f"{headers[i]}: {cell}" if cell is not None else f"{headers[i]}: " for i, cell in enumerate(row))+")")
    
    return "\n".join(extracted_text)

def read_csv(content):
    extracted_text = []
    csv_file = StringIO(content.getvalue().decode('utf-8'))
    reader = csv.reader(csv_file)
    
    rows = list(reader)
    if not rows:
        return ""
    
    headers = rows[0]  # First row as headers
    for row in rows[1:]:  # Iterate over remaining rows
        extracted_text.append("("+",\t".join(f"{headers[i]}: {cell}" if cell else f"{headers[i]}: " for i, cell in enumerate(row))+")")
    
    return "\n".join(extracted_text)


@app.route('/extract_text', methods=['POST'])
def extract_text():
    try:
        data = request.get_json()
        file_url = data.get('file_url')
        if not file_url:
            return jsonify({'error': 'No file URL provided'}), 400
        
        file_type = get_file_type(file_url)
        if file_type not in valid_file_types:
            return jsonify({'error': 'Unsupported file type'}), 400
        
        response = requests.get(file_url, stream=True)
        response.raise_for_status()

        file_size = int(response.headers.get('Content-Length', 0))  # file size in bytes
        if (file_size > 50000000): #50mb limit
            return jsonify({'error': 'File too large'}), 413
        content = BytesIO(response.content)

        print(f"Parsed {file_url}, Size: {file_size} bytes")

        if file_type == "docx":
            extracted_text = read_docx(content)
        elif file_type == "xlsx":
            extracted_text = read_xlsx(content)
        elif file_type == "csv":
            extracted_text = read_csv(content)
        else:
            extracted_text = read_with_PyMuPDF(content, file_type)

        return jsonify({'extracted_text': extracted_text, 'file_size': file_size}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route("/", methods=['GET'])
def default_route():
    return jsonify({'message': "Parser API is working"}), 200

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8081)
